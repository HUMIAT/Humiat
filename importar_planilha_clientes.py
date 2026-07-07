"""
Migração única da planilha Controle_Cliente.xlsm para o Organiza.

Uso:
    1. Coloque o arquivo Controle_Cliente.xlsm na raiz do projeto.
    2. Execute: python importar_planilha_clientes.py

O script:
    - executa somente uma vez por banco de dados;
    - grava uma trava na tabela controle_importacao;
    - limpa clientes/equipamentos somente antes da primeira importação;
    - lê a aba CADASTRO;
    - cria clientes apenas a partir da primeira planilha;
    - lê a aba EQUIPAMENTOS;
    - cria equipamentos somente quando existir vínculo com cliente já cadastrado;
    - ignora equipamentos soltos para evitar cadastros duplicados ou sem dono.

Segurança:
    Depois que a importação for concluída, o script não roda novamente no mesmo banco,
    evitando duplicidade ou sobrescrita dos dados atuais.
"""

from datetime import datetime, date
from pathlib import Path
import re

from openpyxl import load_workbook

from database import SessionLocal, engine, Base
from app import Cliente, Equipamento, limpar_telefone
from sqlalchemy import inspect, text as sql_text


ARQUIVO_PADRAO = Path("Controle_Cliente.xlsm")


def texto(valor):
    if valor is None:
        return None
    valor = str(valor).strip()
    if valor in ("", "None"):
        return None
    return valor


def inteiro(valor):
    if valor is None:
        return None
    try:
        return int(float(valor))
    except Exception:
        return None


def data_excel(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    return None


def dinheiro(valor):
    if valor is None:
        return None
    try:
        return str(float(valor)).rstrip("0").rstrip(".")
    except Exception:
        return texto(valor)


def mapa_cabecalho(ws):
    primeira_linha = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return {str(nome).strip().upper(): indice for indice, nome in enumerate(primeira_linha) if nome is not None}


def valor_linha(linha, mapa, coluna):
    indice = mapa.get(coluna.upper())
    if indice is None or indice >= len(linha):
        return None
    return linha[indice]



NOME_IMPORTACAO = "controle_cliente_xlsm_v1"


def garantir_tabela_controle_importacao():
    """Cria a tabela que registra importações já executadas."""
    with engine.begin() as conn:
        conn.execute(sql_text("""
            CREATE TABLE IF NOT EXISTS controle_importacao (
                nome VARCHAR(120) PRIMARY KEY,
                arquivo VARCHAR(255),
                executada_em TIMESTAMP,
                clientes_criados INTEGER DEFAULT 0,
                clientes_atualizados INTEGER DEFAULT 0,
                equipamentos_criados INTEGER DEFAULT 0,
                equipamentos_ignorados INTEGER DEFAULT 0,
                equipamentos_sem_cliente INTEGER DEFAULT 0
            )
        """))


def importacao_ja_executada(nome=NOME_IMPORTACAO):
    """Retorna True quando esta importação já foi concluída neste banco."""
    garantir_tabela_controle_importacao()
    with engine.begin() as conn:
        resultado = conn.execute(
            sql_text("SELECT nome FROM controle_importacao WHERE nome = :nome"),
            {"nome": nome},
        ).first()
    return resultado is not None


def registrar_importacao_concluida(
    arquivo,
    clientes_criados,
    clientes_atualizados,
    equipamentos_criados,
    equipamentos_ignorados,
    equipamentos_sem_cliente,
    nome=NOME_IMPORTACAO,
):
    """Marca a importação como concluída para impedir nova execução."""
    garantir_tabela_controle_importacao()
    with engine.begin() as conn:
        conn.execute(
            sql_text("""
                INSERT INTO controle_importacao (
                    nome,
                    arquivo,
                    executada_em,
                    clientes_criados,
                    clientes_atualizados,
                    equipamentos_criados,
                    equipamentos_ignorados,
                    equipamentos_sem_cliente
                )
                VALUES (
                    :nome,
                    :arquivo,
                    :executada_em,
                    :clientes_criados,
                    :clientes_atualizados,
                    :equipamentos_criados,
                    :equipamentos_ignorados,
                    :equipamentos_sem_cliente
                )
            """),
            {
                "nome": nome,
                "arquivo": str(arquivo),
                "executada_em": datetime.now(),
                "clientes_criados": clientes_criados,
                "clientes_atualizados": clientes_atualizados,
                "equipamentos_criados": equipamentos_criados,
                "equipamentos_ignorados": equipamentos_ignorados,
                "equipamentos_sem_cliente": equipamentos_sem_cliente,
            },
        )


def garantir_colunas_novas():
    """Garante que bancos antigos tenham os campos novos antes da migração."""
    insp = inspect(engine)
    with engine.begin() as conn:
        tabelas = insp.get_table_names()
        if "clientes" in tabelas:
            colunas = [c["name"] for c in insp.get_columns("clientes")]
            novas = {
                "documento": "VARCHAR(30)",
                "cep": "VARCHAR(20)",
                "cidade": "VARCHAR(120)",
                "bairro": "VARCHAR(120)",
                "endereco": "VARCHAR(255)",
                "email": "VARCHAR(140)",
                "pacote": "VARCHAR(30)",
                "falta_pacote": "INTEGER",
                "plano": "VARCHAR(60)",
            }
            for coluna, tipo in novas.items():
                if coluna not in colunas:
                    conn.execute(sql_text(f"ALTER TABLE clientes ADD COLUMN {coluna} {tipo}"))

def importar(caminho=ARQUIVO_PADRAO):
    caminho = Path(caminho)
    if not caminho.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")

    Base.metadata.create_all(bind=engine)
    garantir_colunas_novas()
    garantir_tabela_controle_importacao()

    if importacao_ja_executada():
        print("Importação bloqueada.")
        print("A planilha Controle_Cliente.xlsm já foi importada neste banco.")
        print("Nenhum cliente ou equipamento foi alterado.")
        return

    wb = load_workbook(caminho, read_only=True, data_only=True, keep_vba=True)
    db = SessionLocal()

    criados = atualizados = equipamentos_criados = equipamentos_ignorados = sem_cliente = 0
    clientes_por_id = {}
    clientes_por_telefone = {}
    clientes_por_nome = {}

    try:
        # Migração única: deixa o banco limpo para refletir exatamente a planilha atual.
        db.query(Equipamento).delete()
        db.query(Cliente).delete()
        db.commit()

        ws = wb["CADASTRO"]
        cab = mapa_cabecalho(ws)

        for linha in ws.iter_rows(min_row=2, values_only=True):
            nome = texto(valor_linha(linha, cab, "Nome no WhatsApp"))
            telefone = limpar_telefone(texto(valor_linha(linha, cab, "Telefone")) or "")
            if not nome or not telefone:
                continue

            cliente = clientes_por_telefone.get(telefone)
            if not cliente:
                cliente = db.query(Cliente).filter(Cliente.telefone == telefone).first()
            if not cliente:
                cliente = Cliente(nome=nome, telefone=telefone)
                db.add(cliente)
                criados += 1
            else:
                atualizados += 1

            cliente.nome = nome
            cliente.empresa = texto(valor_linha(linha, cab, "Empresa"))
            cliente.documento = texto(valor_linha(linha, cab, "Documento") or valor_linha(linha, cab, "CPF/CNPJ"))
            cliente.cep = texto(valor_linha(linha, cab, "CEP"))
            cliente.cidade = texto(valor_linha(linha, cab, "Cidade"))
            cliente.bairro = texto(valor_linha(linha, cab, "Bairro"))
            cliente.endereco = texto(valor_linha(linha, cab, "Endereço") or valor_linha(linha, cab, "Endereco"))
            cliente.email = texto(valor_linha(linha, cab, "Email") or valor_linha(linha, cab, "E-mail"))
            cliente.pacote = texto(valor_linha(linha, cab, "PACOTE"))
            cliente.falta_pacote = inteiro(valor_linha(linha, cab, "FALTA PACOTE"))
            cliente.plano = texto(valor_linha(linha, cab, "PLANO"))

            id_cadastro = texto(valor_linha(linha, cab, "ID CADASTRO"))
            if id_cadastro:
                clientes_por_id[id_cadastro] = cliente
            clientes_por_telefone[telefone] = cliente
            if nome:
                clientes_por_nome[nome.strip().lower()] = cliente

        db.commit()

        ws = wb["EQUIPAMENTOS"]
        cab = mapa_cabecalho(ws)

        for linha in ws.iter_rows(min_row=2, values_only=True):
            telefone = limpar_telefone(texto(valor_linha(linha, cab, "TELEFONE")) or "")
            id_cadastro = texto(valor_linha(linha, cab, "ID CADASTRO"))
            nome_cliente = texto(valor_linha(linha, cab, "CLIENTE"))

            cliente = None
            if id_cadastro and id_cadastro in clientes_por_id:
                cliente = clientes_por_id[id_cadastro]
            if not cliente and telefone:
                cliente = db.query(Cliente).filter(Cliente.telefone == telefone).first()
            if not cliente and nome_cliente:
                cliente = clientes_por_nome.get(nome_cliente.strip().lower())

            # Importante:
            # equipamento sem vínculo real com a aba CADASTRO é ignorado.
            # Não criamos cliente a partir da aba EQUIPAMENTOS para evitar redundância.
            if not cliente:
                sem_cliente += 1
                continue

            tipo = texto(valor_linha(linha, cab, "TIPO"))
            modelo = texto(valor_linha(linha, cab, "MODELO"))
            data_compra = data_excel(valor_linha(linha, cab, "DATA COMPRA"))
            valor = dinheiro(valor_linha(linha, cab, "VALOR"))

            existente = db.query(Equipamento).filter(
                Equipamento.cliente_id == cliente.id,
                Equipamento.tipo == tipo,
                Equipamento.modelo == modelo,
                Equipamento.data_compra == data_compra,
                Equipamento.valor == valor,
            ).first()

            if existente:
                equipamentos_ignorados += 1
                continue

            falta = dinheiro(valor_linha(linha, cab, "FALTA"))
            status = "Ativo"
            try:
                if float(falta or 0) > 0:
                    status = "Pendente"
            except Exception:
                pass

            equipamento = Equipamento(
                cliente_id=cliente.id,
                tipo=tipo,
                modelo=modelo,
                pacote=texto(valor_linha(linha, cab, "PACOTE")),
                falta_pacote=inteiro(valor_linha(linha, cab, "PACOTE FALTA")),
                plano=texto(valor_linha(linha, cab, "PLANO")),
                valor=valor,
                pago=dinheiro(valor_linha(linha, cab, "PAGO")),
                falta=falta,
                data_compra=data_compra,
                previsao_entrega=data_excel(valor_linha(linha, cab, "PREVISÃO DE ENTREGA")),
                maquina=texto(valor_linha(linha, cab, "MAQUINA")),
                rede_instalada=texto(valor_linha(linha, cab, "REDE INSTALADA")),
                anydesk=texto(valor_linha(linha, cab, "ANYDESK")),
                status=status,
            )
            db.add(equipamento)
            equipamentos_criados += 1

        db.commit()

        registrar_importacao_concluida(
            arquivo=caminho,
            clientes_criados=criados,
            clientes_atualizados=atualizados,
            equipamentos_criados=equipamentos_criados,
            equipamentos_ignorados=equipamentos_ignorados,
            equipamentos_sem_cliente=sem_cliente,
        )

        print("Migração concluída.")
        print(f"Clientes criados: {criados}")
        print(f"Clientes atualizados: {atualizados}")
        print(f"Equipamentos criados: {equipamentos_criados}")
        print(f"Equipamentos ignorados por duplicidade: {equipamentos_ignorados}")
        print(f"Equipamentos ignorados sem vínculo no cadastro: {sem_cliente}")

    finally:
        db.close()


if __name__ == "__main__":
    importar()
